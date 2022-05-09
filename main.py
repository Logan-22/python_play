from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# To create New sheet
# wb = load_workbook('excel/test_python.xlsx')
# ws = wb['Marks']

# wb.create_sheet("Test")
# wb.save('excel/test_python.xlsx')
# print(wb.sheetnames)

# To create New Workbook

# wb = Workbook()
# ws = wb.active
# ws.title = "Data"

# ws.append(["Loganand", "Is", "Great", "!"])

# wb.save("excel/Test.xlsx")

# To Access Multiple rows

wb = load_workbook("excel/Test.xlsx")
ws = wb.active

for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

wb.save("excel/Test.xlsx")